from __future__ import annotations

from collections import Counter
from datetime import datetime
import json
from pathlib import Path
import re


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT.parent / "deliverables" / "Dossche_Mills_AI_Scorecard_v2.xlsx"
OUTPUT = ROOT / "src" / "data" / "usecases.generated.json"

GROUP_SHEETS = [
    ("Thera", "thera", "Group 1 - Thera"),
    ("Procurement", "procurement", "Group 2 - Procurement"),
    ("R&D / Product", "rd-product", "Group 3 - R&D Product"),
]

GROUP_CONTEXT = {
    "Thera": {
        "diagnosis": "Thera has two layers of challenge: operational complexity in carbon and standards workflows, and commercial difficulty translating Terra into a quantified procurement story.",
        "manual_workflow": "Work spans multiple calculation tools, several data sources beyond SAP, repeated formatting work, and a lot of context still living in people, meeting notes, and scattered files.",
        "primary_users": ["Sustainability team", "Commercial team", "Terra program owners"],
        "default_sources": [
            "Terra program files",
            "Customer sustainability questionnaires",
            "Carbon calculation tools",
            "Meeting notes and CRM context",
            "Recipe and site data",
            "Standards and regulatory updates",
        ],
    },
    "Procurement": {
        "diagnosis": "Procurement's main issue is not lack of data, but reconciliation, harmonization, and delayed visibility across SAP, Archiva, Excel, scorecards, emails, and local workarounds.",
        "manual_workflow": "Many decisions depend on manually reconciling forecast, contracts, actuals, supplier signals, and planning context before any action can be taken.",
        "primary_users": ["Procurement managers", "Planners", "Buyers", "Category owners"],
        "default_sources": [
            "SAP extracts",
            "Archiva and local Excel trackers",
            "Contracts and tender files",
            "Supplier emails and meeting notes",
            "Market and commodity signals",
            "Planning and actuals files",
        ],
    },
    "R&D / Product": {
        "diagnosis": "R&D / Product has a foundational knowledge retrieval problem: information is fragmented across reports, formulas, emails, SharePoint, CRM, application files, and master-data workflows.",
        "manual_workflow": "Teams repeatedly search for past trials, similar formulas, customer context, and technical truth through disconnected files before they can respond or create something new.",
        "primary_users": ["R&D team", "Product management", "Application specialists", "Technical support"],
        "default_sources": [
            "Trial reports and test files",
            "Product specs and formula libraries",
            "CRM and customer history",
            "Emails and meeting notes",
            "Technical sheets and brochures",
            "Master-data and import templates",
        ],
    },
}

THEME_KEYWORDS = {
    "knowledge": ["knowledge", "taxonomy", "tagging", "history", "search", "retrieval", "radar"],
    "carbon": ["carbon", "footprint", "terra", "wheat-level", "traceability"],
    "commercial": ["customer", "sales", "proposal", "segmentation", "business case", "campaign", "account"],
    "monitoring": ["monitor", "standards", "legislation", "regulation", "radar"],
    "procurement-visibility": ["dashboard", "gap", "forecast", "actuals", "coverage", "visibility", "material-level"],
    "contracts-tenders": ["contract", "supplier", "tender", "ranking"],
    "market-intelligence": ["market", "competitor", "price", "pricing", "energy", "intelligence"],
    "logistics": ["vessel", "barge", "eta", "shipment", "route", "loading"],
    "product-creation": ["copy-formula", "product", "master-data", "creation", "workflow"],
    "optimization": ["optimization", "scenario", "mix", "counter-typing", "similarity", "solution finder"],
    "quality-compliance": ["proofreading", "complaint", "issue", "compliance", "technical"],
    "admin-relief": ["email", "meeting", "capture", "time-registration", "onboarding"],
}

KEYWORD_SOURCES = {
    "questionnaire": ["Questionnaire templates", "Prior customer responses", "Sustainability evidence packs"],
    "survey": ["Questionnaire templates", "Prior customer responses", "Sustainability evidence packs"],
    "knowledge": ["SOPs", "Expert notes", "FAQs", "Business rules and assumptions"],
    "standards": ["Regulatory feeds", "Standards publications", "Internal interpretation notes"],
    "legislation": ["Regulatory feeds", "Standards publications", "Internal interpretation notes"],
    "carbon": ["Recipe data", "Packaging data", "Transport data", "Production and site data"],
    "footprint": ["Recipe data", "Packaging data", "Transport data", "Production and site data"],
    "meeting": ["Meeting transcripts", "Action trackers", "CRM notes", "Calendar events"],
    "crm": ["CRM records", "Meeting transcripts", "Account notes"],
    "forecast": ["Forecast files", "Sales input", "Contracts", "Actuals"],
    "dashboard": ["SAP exports", "Operational scorecards", "Excel trackers", "Actuals files"],
    "tender": ["RFX files", "Supplier responses", "Evaluation criteria", "Pricing tables"],
    "contract": ["Contract library", "Standard clauses", "Review checklist", "Renewal dates"],
    "supplier": ["Supplier scorecards", "Supplier emails", "Meeting notes", "Contracts"],
    "eta": ["Shipment status", "Logistics schedules", "Route data", "Load planning"],
    "barge": ["Shipment status", "Logistics schedules", "Route data", "Load planning"],
    "price": ["Supplier quotes", "Market feeds", "Historical pricing", "Meeting intelligence"],
    "energy": ["Energy market feeds", "Demand assumptions", "Consumption data", "Buy history"],
    "formula": ["Formula database", "Ingredient specs", "Past trials", "Application notes"],
    "proofreading": ["Technical sheets", "Brochures", "Master data", "Approved product truth"],
    "complaint": ["Complaint logs", "Customer emails", "Timeline notes", "Quality records"],
    "email": ["Inbox metadata", "Message bodies", "Shared folders", "Task cues"],
    "time-registration": ["Calendar data", "Email activity", "Chat signals", "Time-entry tool"],
    "radar": ["Retailer websites", "Competitor signals", "Launch tracking", "Category observations"],
}

DELIVERY_MODE_OVERRIDES = {
    "T1": "WITH You",
    "T2": "WITH You",
    "T4": "FOR You",
    "T8": "FOR You",
    "T10": "FOR You",
    "T11": "WITH You",
    "T12": "FOR You",
    "T13": "WITH You",
    "T14": "BY You",
    "T15": "WITH You",
    "T16": "BY You",
    "T17": "WITH You",
    "T18": "WITH You",
    "T19": "WITH You",
    "P1": "FOR You",
    "P2": "WITH You",
    "P3": "WITH You",
    "P4": "WITH You",
    "P5": "FOR You",
    "P6": "WITH You",
    "P7": "FOR You",
    "P8": "FOR You",
    "P9": "WITH You",
    "P10": "WITH You",
    "P12": "BY You",
    "P14": "WITH You",
    "P15": "WITH You",
    "P16": "WITH You",
    "P17": "WITH You",
    "P18": "BY You",
    "P20": "WITH You",
    "P21": "BY You",
    "R1": "FOR You",
    "R2": "WITH You",
    "R3": "WITH You",
    "R4": "FOR You",
    "R5": "FOR You",
    "R6": "FOR You",
    "R7": "FOR You",
    "R8": "BY You",
    "R11": "WITH You",
    "R12": "WITH You",
    "R13": "FOR You",
    "R14": "FOR You",
    "R15": "FOR You",
    "R16": "FOR You",
    "R17": "BY You",
    "R19": "BY You",
    "R20": "BY You",
    "R21": "WITH You",
}

CASE_OVERRIDES = {
    "R15": {
        "description": "Build a bounded replacement for part of Innova by scraping selected retailers, extracting rich product attributes, classifying them into an Innova-like taxonomy, and exporting a recurring tracker for priority categories.",
        "evidenceNotes": "The real need is not generic scraping, but replacing part of Innova with a structured, filterable market intelligence layer for West Europe and priority categories.",
        "kpi": "Publish a monthly or quarterly tracker for 2-3 priority categories in West Europe, with normalized product rows and inferred launch/package events.",
        "overview": "R15 is not a generic chatbot or scraping bot. It is a bounded partial substitute for Innova: retailer discovery, structured extraction, taxonomy-based classification, and Excel-friendly recurring market tracking for priority categories.",
        "problem": "The team is losing access to Innova, a paid market-intelligence source they use to understand launches, market structure, category evolution, claims, formats, and retailer assortment by submarket. Without a replacement, they lose a structured lens on what exists in the market and how it is changing.",
        "themes": ["market-intelligence", "monitoring", "knowledge"],
        "primaryUsers": ["Iris", "Product management", "R&D team", "Application specialists"],
        "currentWorkflow": [
            "Today the team relies on Innova exports and its filtering logic to analyze categories, subcategories, countries, positioning, texture, event type, shelving, packaging, ingredients, flavors, formats, and nutrition.",
            "The working unit is not a webpage but a normalized product row that can be filtered in Excel-style analysis, including retailer, event date, category, subcategory, claims, positioning, packaging, price, ingredients, and nutrition.",
            "As the Innova subscription is going away, they would otherwise have to reconstruct this market view manually from retailer sites, which is far slower and much less comparable across countries and categories.",
        ],
        "solutionConcept": [
            "Build a retailer observation pipeline for selected West Europe retailers and a bounded set of categories such as cakes and coated poultry.",
            "Extract product-page text, images, packaging information, claims, ingredients, nutrition, shelving, pricing, and retailer metadata, then normalize it into a row structure comparable to the sample Innova export.",
            "Add an Innova-like classification layer for category, subcategory, region, country, positioning, texture, shelving, and event so the output is analytically useful rather than just a raw scrape.",
            "Use recurring snapshots plus human review to infer events such as new product, new package, import, reformulation, or shelf snapshot where feasible, instead of promising perfect parity with Innova in phase 1.",
        ],
        "whyItMatters": [
            "This directly addresses the budget-driven loss of Innova while preserving a meaningful share of its business value for market and category intelligence.",
            "The team uses this intelligence to size markets, segment submarkets, compare offer structures, and answer questions such as what share of a category is coated, which formats dominate, and which claims or positioning patterns repeat.",
            "The first useful output is not a chatbot answer but a living, exportable dataset that product managers can filter by category, country, retailer, positioning, event, texture, and other dimensions.",
            "The success metric already proposed for this use case is: Publish a monthly or quarterly tracker for 2-3 priority categories in West Europe, with normalized product rows and inferred launch/package events.",
        ],
        "dataSources": [
            "Retailer category pages and assortment listings",
            "Retailer product detail pages",
            "Front and back product images where available",
            "Historical crawl snapshots from previous runs",
            "Client-provided Innova exports and sample Excel outputs",
            "Innova-style taxonomy for category, subcategory, region, country, positioning, texture, event, and shelving",
            "Claim, ingredient, packaging, and nutrition extraction rules",
            "Country and retailer scope configuration for West Europe",
        ],
        "dependencies": [
            "Agreement on the first 2-3 priority categories and the first retailer set",
            "A bounded taxonomy that mirrors the Innova filters the business actually uses, rather than every possible dimension from day one",
            "Rules for when the system may label a product as new product, new package, import, reformulation, or only a shelf snapshot",
            "A human validation step for classification quality, especially across countries, languages, and private-label ranges",
            "A storage layer that preserves recurring snapshots so observations can be compared over time",
        ],
        "risks": [
            "Retailer sites vary widely in structure and richness, so extracted fields will not be equally available across every source.",
            "Event detection from snapshots is an inference problem, not a perfect truth source, especially for reformulation and new package logic.",
            "Trying to replicate full Innova breadth in phase 1 would turn a bounded Quick Win into a much larger data-platform project.",
            "Without a carefully scoped taxonomy, the scrape could generate data volume without preserving the analytical usefulness that Innova provided.",
        ],
        "evidence": [
            "Stefanie explicitly framed the request as compensating for Innova by scraping retailer websites and shared Innova exports plus screenshots of the filters they use.",
            "The screenshots act as a functional specification: food category and subcategory, country and region, positioning, texture, event, shelving, and additional attribute filters are all part of the desired experience.",
            "The sample export shows the required output shape is a normalized table with retailer, event, category, brand, product name, claims, ingredients, nutrition, price, packaging, and other structured fields.",
            "Internally the team clarified that the goal is not a conversational bot first, but a recurring market-intelligence data layer that can partially replace a EUR 20k/year source.",
        ],
        "implementationNotes": [
            "Recommended delivery mode: WITH You.",
            "Recommended portfolio position: First Wave.",
            "Scope the MVP to selected West Europe retailers, 2-3 priority categories, and an exportable dataset before adding a conversational interface.",
            "Treat Innova parity as out of scope for phase 1; the target is a useful partial substitute, not a one-to-one clone.",
        ],
        "scoreSummary": "The current Quick Win scoring only makes sense for a bounded MVP: selected West Europe retailers, 2-3 categories, structured extraction, taxonomy, and recurring export. A full Innova replacement across categories, countries, and event logic would be materially more complex than this score suggests.",
    },
    "R16": {
        "description": "Track retailer assortment persistence over time so the team can infer shelf survival, disappearance, and product longevity instead of relying only on launch counts.",
        "evidenceNotes": "The team explicitly concluded that launch tracking alone is not enough; a better signal is whether products stay visible in retailer assortments over time.",
        "kpi": "Refresh recurring snapshots for 2-3 priority categories and report product persistence, disappearance, and inferred shelf survival by retailer and market.",
        "overview": "R16 is the longitudinal layer on top of R15. Instead of only asking what is new, it asks which products remain present, which disappear, and how long ranges survive in visible retail assortment.",
        "problem": "A launch database can show novelty, but it does not answer whether products survive in market, remain listed by retailers, or disappear after a short period. The team explicitly challenged launch-counting as an incomplete proxy for commercial reality.",
        "themes": ["market-intelligence", "monitoring", "knowledge"],
        "primaryUsers": ["Iris", "Product management", "R&D team", "Commercial intelligence users"],
        "currentWorkflow": [
            "Today most market-intelligence tools emphasize launches and product events, but that still leaves the team blind to whether products persist in retailer assortment and for how long.",
            "To answer survival questions, the same categories and retailers would need to be revisited on a recurring basis and matched product by product across snapshots.",
            "Without that recurring layer, the team can see that something appeared once, but not whether it remained, expanded, disappeared, or changed packaging over time.",
        ],
        "solutionConcept": [
            "Build a recurring snapshot process on top of the retailer-observation layer so the same categories, retailers, and products are revisited monthly or quarterly.",
            "Match products across time using product identity signals such as retailer, brand, title, pack size, images, barcode, and other stable attributes, then infer persistence or disappearance.",
            "Calculate shelf-survival style metrics such as first seen, last seen, continuous presence window, assortment churn, and category-level persistence by retailer or country.",
            "Use the same Innova-like taxonomy from R15 so longevity can be analyzed by category, subcategory, positioning, texture, shelving, and other filters that matter to the business.",
        ],
        "whyItMatters": [
            "This gives the team a more meaningful signal than launches alone, because market reality is not just who launched, but what stayed listed and visible over time.",
            "A survival view helps separate novelty from actual staying power and turns recurring retailer scraping into a longitudinal market signal rather than a one-off snapshot.",
            "This becomes especially valuable once the team wants to compare retailer ranges, identify category churn, and observe which formats or claims have durable shelf presence.",
            "The success metric already proposed for this use case is: Refresh recurring snapshots for 2-3 priority categories and report product persistence, disappearance, and inferred shelf survival by retailer and market.",
        ],
        "dataSources": [
            "Historical retailer crawl snapshots",
            "Retailer category pages and assortment listings",
            "Retailer product detail pages",
            "Front and back product images where available",
            "Identity and matching fields such as brand, title, barcode, pack size, and retailer",
            "The taxonomy layer from R15 for category, subcategory, positioning, texture, event, and shelving",
            "Manual validation sets for product matching and event interpretation",
            "Category and retailer scope definitions for recurring tracking",
        ],
        "dependencies": [
            "R15 or an equivalent retailer-observation layer needs to exist first, because longevity requires recurring snapshots of a structured base dataset",
            "Agreed rules for what counts as persistence, disappearance, relaunch, new package, or reformulation",
            "A reliable product-matching approach across time despite packaging changes, retailer naming changes, and incomplete fields",
            "A recurring crawl cadence and storage model so the same retailers and categories are revisited consistently",
            "Human review for ambiguous product identity or event changes in the early phases",
        ],
        "risks": [
            "Longevity logic is only as good as the recurring snapshot discipline and the ability to match the same product across time.",
            "A product disappearing from a retailer website does not always mean it failed commercially; it may reflect assortment rotation, page changes, or temporary listing gaps.",
            "If R15 taxonomy and extraction quality are weak, the survival layer will inherit that noise and produce unstable signals.",
            "This quickly becomes more complex than launch tracking because it requires history, identity resolution, and event interpretation rather than single-date extraction.",
        ],
        "evidence": [
            "In the internal discussion, the team explicitly challenged launches as an imperfect metric and said it would be more valuable to capture the full observable catalogue and revisit it periodically.",
            "The proposed better signal was persistence in assortment: whether products are still visible after repeated retailer observations, not just whether they once appeared as new.",
            "This is why R16 is not a duplicate of R15 but a strategic next layer: recurring observation, product matching, and longitudinal interpretation.",
            "The Innova screenshots and export structure still matter here because longevity only becomes useful if it can be sliced by the same business dimensions the team already uses.",
        ],
        "implementationNotes": [
            "Recommended delivery mode: WITH You.",
            "Recommended portfolio position: Keep Visible.",
            "Sequence this after the bounded R15 MVP, because shelf survival depends on recurring structured snapshots rather than one-off scraping.",
            "Start with a limited cadence and limited category set so persistence logic can be validated before expanding coverage.",
        ],
        "scoreSummary": "This remains a Strategic Bet because the value is clear but the hard part is not scraping once; it is maintaining recurring snapshots, matching products across time, and interpreting persistence versus disappearance reliably.",
    },
}


def slugify(value: str) -> str:
    value = value.lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def clean(value):
    return value.strip() if isinstance(value, str) else value


def normalize_date(value) -> str | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


def dedupe(items: list[str]) -> list[str]:
    seen = set()
    output: list[str] = []
    for item in items:
        if item and item not in seen:
            output.append(item)
            seen.add(item)
    return output


def text_blob(case: dict) -> str:
    return " ".join(
        [
            case["title"],
            case["description"],
            case["evidenceNotes"],
            case["kpi"],
            case["filterLens"],
            case["quadrant"],
        ]
    ).lower()


def detect_themes(case: dict) -> list[str]:
    blob = text_blob(case)
    themes = [theme for theme, keywords in THEME_KEYWORDS.items() if any(keyword in blob for keyword in keywords)]
    return themes or ["general"]


def infer_data_sources(case: dict) -> list[str]:
    sources = list(GROUP_CONTEXT[case["departmentLabel"]]["default_sources"])
    blob = text_blob(case)
    for keyword, additions in KEYWORD_SOURCES.items():
        if keyword in blob:
            sources.extend(additions)
    return dedupe(sources)[:10]


def infer_current_workflow(case: dict) -> list[str]:
    lines = [
        GROUP_CONTEXT[case["departmentLabel"]]["manual_workflow"],
        f"This specific workflow is currently handled through manual coordination around {case['description'].lower()}",
    ]
    if case["quadrant"] == "Optimize":
        lines.append("The team is spending time translating raw inputs into a fixed output shape before anyone can act on the result.")
    elif case["quadrant"] == "Insights":
        lines.append("Decision quality depends on assembling partial signals fast enough to influence the next commercial or operational choice.")
    else:
        lines.append("The work relies on expert interpretation, but the supporting information is not yet packaged into a reusable operating layer.")
    return lines


def infer_solution_concept(case: dict) -> list[str]:
    lines = [
        case["description"],
        f"The solution should be delivered as a {case['deliveryMode']} capability with a clear human-in-the-loop review step where business judgment still matters.",
    ]
    if case["quadrant"] == "Optimize":
        lines.append("A good first release should reduce repetitive manual preparation while preserving traceability of the output.")
    elif case["quadrant"] == "Insights":
        lines.append("A good first release should surface decision-ready context, not just a data dump, and explain what changed and why it matters.")
    else:
        lines.append("A good first release should augment experts with reusable context and recommended next steps instead of replacing judgment.")
    return lines


def infer_why_it_matters(case: dict) -> list[str]:
    return [
        case["evidenceNotes"],
        f"The success metric already proposed for this use case is: {case['kpi']}.",
        f"In the current scorecard it sits as a {case['priority']} with {case['aiMaturity']} AI maturity and a first visible win date of {case['firstVisibleWinDate']}.",
    ]


def infer_dependencies(case: dict, data_sources: list[str]) -> list[str]:
    deps = [
        "Clear source ownership and access for the required inputs",
        "An agreed pilot scope with one accountable business owner",
    ]
    if case["complexityScore"] >= 6:
        deps.append("A defined integration and architecture plan across the affected systems")
    if case["quadrant"] in {"Insights", "Enhance"}:
        deps.append("Business rules for how recommendations should be interpreted and acted on")
    if "knowledge" in case["themes"] or "product-creation" in case["themes"]:
        deps.append("A minimum taxonomy or structure so the content can be indexed consistently")
    if any(source in data_sources for source in ["Regulatory feeds", "Standards publications"]):
        deps.append("An approved source list for compliance-sensitive monitoring")
    return dedupe(deps)


def infer_risks(case: dict) -> list[str]:
    risks = ["Fragmented source data could limit early precision and user trust if not scoped carefully."]
    if case["complexityScore"] >= 6:
        risks.append("Integration and change-management effort may dominate the first phase if too much scope is included.")
    if case["aiMaturityValue"] < 60:
        risks.append("The use case should not be oversold before the enabling structure is in place.")
    if case["quadrant"] == "Insights":
        risks.append("Users may still default to gut feel unless the output is timely, explainable, and embedded in the actual decision moment.")
    if case["quadrant"] == "Optimize":
        risks.append("A pilot can fail if it automates the wrong slice of the workflow and leaves the real bottleneck untouched.")
    return dedupe(risks)


def infer_evidence(case: dict) -> list[str]:
    return [
        case["evidenceNotes"],
        GROUP_CONTEXT[case["departmentLabel"]]["diagnosis"],
        f"The AI Compass lens for this use case is {case['filterLens']} and the recommended motion is {case['recommendation']}.",
    ]


def score_summary(case: dict) -> str:
    return (
        f"Value {case['valueScore']} driven by Time {case['valueDrivers']['time']}, Cost {case['valueDrivers']['cost']}, "
        f"Criticality {case['valueDrivers']['criticality']}, and Friction {case['valueDrivers']['pain']}. "
        f"Complexity {case['complexityScore']} driven by Delivery Time {case['complexityDrivers']['time']}, "
        f"Budget {case['complexityDrivers']['budget']}, Expertise {case['complexityDrivers']['expertise']}, "
        f"and Systems {case['complexityDrivers']['systems']}."
    )


def implementation_notes(case: dict) -> list[str]:
    notes = [
        f"Recommended delivery mode: {case['deliveryMode']}.",
        f"Recommended portfolio position: {case['recommendation']}.",
    ]
    if case["priority"] == "Quick Win":
        notes.append("This should be shaped as a bounded pilot with visible adoption in a single team or workflow first.")
    elif case["priority"] == "Strategic Bet":
        notes.append("This should be sequenced behind enabling data, taxonomy, or integration work so the pilot is not architecture-blind.")
    else:
        notes.append("This should stay visible as a lower-priority or later-phase item rather than compete with stronger first-wave opportunities.")
    return notes


def apply_case_overrides(case: dict) -> dict:
    overrides = CASE_OVERRIDES.get(case["code"])
    if not overrides:
        return case

    for key, value in overrides.items():
        case[key] = value

    return case


def apply_delivery_model_override(case: dict) -> dict:
    case["deliveryMode"] = DELIVERY_MODE_OVERRIDES.get(case["code"], case["deliveryMode"])
    return case


def collect_cases() -> tuple[list[dict], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(WORKBOOK)
    cases: list[dict] = []
    for department_label, department_slug, sheet_name in GROUP_SHEETS:
        ws = wb[sheet_name]
        for row in range(3, ws.max_row + 1):
            code = clean(ws.cell(row, 1).value)
            if not code:
                continue
            case = {
                "type": "core",
                "code": code,
                "slug": slugify(f"{code}-{clean(ws.cell(row, 2).value)}"),
                "title": clean(ws.cell(row, 2).value),
                "departmentLabel": department_label,
                "departmentSlug": department_slug,
                "brace": clean(ws.cell(row, 3).value),
                "quadrant": clean(ws.cell(row, 4).value),
                "filterLens": clean(ws.cell(row, 5).value),
                "valueDrivers": {
                    "time": int(ws.cell(row, 6).value),
                    "cost": int(ws.cell(row, 7).value),
                    "criticality": int(ws.cell(row, 8).value),
                    "pain": int(ws.cell(row, 9).value),
                },
                "valueScore": int(ws.cell(row, 10).value),
                "complexityDrivers": {
                    "time": int(ws.cell(row, 11).value),
                    "budget": int(ws.cell(row, 12).value),
                    "expertise": int(ws.cell(row, 13).value),
                    "systems": int(ws.cell(row, 14).value),
                },
                "complexityScore": float(ws.cell(row, 15).value),
                "priority": clean(ws.cell(row, 16).value),
                "deliveryMode": clean(ws.cell(row, 17).value),
                "recommendation": clean(ws.cell(row, 18).value),
                "owner": clean(ws.cell(row, 19).value),
                "kpi": clean(ws.cell(row, 20).value),
                "firstVisibleWinDate": normalize_date(ws.cell(row, 21).value),
                "aiMaturity": clean(ws.cell(row, 22).value),
                "aiMaturityValue": int(str(clean(ws.cell(row, 22).value)).replace("%", "")),
                "description": clean(ws.cell(row, 23).value),
                "evidenceNotes": clean(ws.cell(row, 24).value),
            }
            case = apply_delivery_model_override(case)
            case["overview"] = (
                f"{case['title']} is a {case['priority']} {case['quadrant'].lower()} use case for {department_label}. "
                f"It is meant to {case['description'].lower()}"
            )
            case["themes"] = detect_themes(case)
            case["dataSources"] = infer_data_sources(case)
            case["primaryUsers"] = dedupe([case["owner"]] + GROUP_CONTEXT[department_label]["primary_users"])
            case["problem"] = GROUP_CONTEXT[department_label]["diagnosis"]
            case["currentWorkflow"] = infer_current_workflow(case)
            case["solutionConcept"] = infer_solution_concept(case)
            case["whyItMatters"] = infer_why_it_matters(case)
            case["dependencies"] = infer_dependencies(case, case["dataSources"])
            case["risks"] = infer_risks(case)
            case["evidence"] = infer_evidence(case)
            case["implementationNotes"] = implementation_notes(case)
            case["scoreSummary"] = score_summary(case)
            case = apply_case_overrides(case)
            cases.append(case)

    review_ws = wb["Reviewed Add-Ons"]
    add_ons: list[dict] = []
    for row in range(3, review_ws.max_row + 1):
        title = clean(review_ws.cell(row, 2).value)
        if not title:
            continue
        department_label = clean(review_ws.cell(row, 1).value)
        add_ons.append(
            {
                "type": "add-on",
                "slug": slugify(title),
                "title": title,
                "departmentLabel": department_label,
                "departmentSlug": slugify(department_label),
                "keepAs": clean(review_ws.cell(row, 3).value),
                "mappedTo": [part.strip() for part in str(clean(review_ws.cell(row, 4).value)).split(",")],
                "whyKeep": clean(review_ws.cell(row, 5).value),
                "whyNotMain": clean(review_ws.cell(row, 6).value),
                "overview": f"{title} was reviewed from the external draft and retained as a supporting idea rather than a core portfolio line.",
            }
        )
    return cases, add_ons


def attach_relationships(cases: list[dict], add_ons: list[dict]) -> None:
    case_lookup = {case["code"]: case for case in cases}
    for case in cases:
        related = []
        for candidate in cases:
            if candidate["code"] == case["code"]:
                continue
            score = 0
            if candidate["departmentLabel"] == case["departmentLabel"]:
                score += 2
            score += len(set(candidate["themes"]) & set(case["themes"]))
            if candidate["priority"] == case["priority"]:
                score += 1
            if score > 2:
                related.append((score, candidate["code"]))
        related.sort(key=lambda item: (-item[0], item[1]))
        case["relatedUseCases"] = [code for _, code in related[:4]]

    for add_on in add_ons:
        add_on["mappedUseCases"] = [case_lookup[code]["slug"] for code in add_on["mappedTo"] if code in case_lookup]


def build_stats(cases: list[dict], add_ons: list[dict]) -> dict:
    return {
        "coreCaseCount": len(cases),
        "addOnCount": len(add_ons),
        "byDepartment": dict(Counter(case["departmentLabel"] for case in cases)),
        "byPriority": dict(Counter(case["priority"] for case in cases)),
        "byFilter": dict(Counter(case["filterLens"] for case in cases)),
        "byRecommendation": dict(Counter(case["recommendation"] for case in cases)),
    }


def main() -> None:
    try:
        cases, add_ons = collect_cases()
    except ModuleNotFoundError as exc:
        if exc.name != "openpyxl":
            raise
        if OUTPUT.exists():
            workbook_is_newer = WORKBOOK.exists() and WORKBOOK.stat().st_mtime > OUTPUT.stat().st_mtime
            print("openpyxl is not installed; using the existing generated JSON.")
            if workbook_is_newer:
                print("Warning: the workbook is newer than the generated JSON, so the site may be slightly stale.")
            print(f"Using {OUTPUT}")
            return
        raise SystemExit(
            "openpyxl is required to generate the site data from the workbook.\n"
            "Install it in your active Python environment with:\n"
            "python3 -m pip install openpyxl"
        )

    attach_relationships(cases, add_ons)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedAt": datetime.utcnow().isoformat() + "Z",
        "sourceWorkbook": str(WORKBOOK.relative_to(ROOT.parent)),
        "groups": [
            {"label": label, "slug": slug, "diagnosis": GROUP_CONTEXT[label]["diagnosis"]}
            for label, slug, _ in GROUP_SHEETS
        ],
        "priorities": ["Quick Win", "Strategic Bet", "Fill-In", "Trap"],
        "stats": build_stats(cases, add_ons),
        "coreCases": cases,
        "addOns": add_ons,
    }
    OUTPUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
